import os  # Importa el módulo os para interactuar con el sistema operativo
import re
import shutil
import sys
import pandas as pd  # Importa pandas para la manipulación de datos
import openpyxl  # Importa openpyxl para trabajar con archivos Excel
from openpyxl.utils.dataframe import dataframe_to_rows  # Función para convertir un DataFrame de pandas en filas para Excel
from openpyxl.styles import Font  # Importa Font para estilizar las celdas de Excel
from openpyxl.drawing.image import Image  # Importa Image para añadir imágenes a los archivos Excel

ARCHIVO_ESPECIFICACIONES_EXCEL = 'Especificaciones_SP2024.xlsx'  # Define el nombre del archivo de especificaciones
DEPENDENCIAS_BD_Y_PLANTILLAS='bdpl'
# Bloque de código para definir rutas de archivos y directorios, y ejecutar las funciones definidas anteriormente

def get_resource_path():
    """ Retorna la ruta absoluta al recurso, para uso en desarrollo o en el ejecutable empaquetado. """
    if getattr(sys, 'frozen', False):
        # Si el programa ha sido empaquetado, el directorio base es el que PyInstaller proporciona
        base_path = sys._MEIPASS
    else:
        # Si estamos en desarrollo, utiliza la ubicación del script
        base_path = os.path.dirname(os.path.realpath(__file__))

    return base_path

script_directory = get_resource_path()

def crear_carpeta_y_archivos(nueva_carpeta,rutaPdf,rutaLinea):
    
    ruta_nueva_carpeta=os.path.join(script_directory,rutaLinea,nueva_carpeta)


    # Crear la nueva carpeta y subcarpetas
    os.makedirs(ruta_nueva_carpeta, exist_ok=True)
    os.makedirs(os.path.join(ruta_nueva_carpeta, "COT"), exist_ok=True)
    os.makedirs(os.path.join(ruta_nueva_carpeta, "FICHAS_TECNICAS"), exist_ok=True)

    # Copiar y renombrar las plantillas SP y OFERTA
    plantillas = [("SP.xlsx", "SP"), ("OFERTA.xlsx", "OFERTA")]
    for plantilla, nombre in plantillas:
        destino = os.path.join(ruta_nueva_carpeta, f"{nombre} {nueva_carpeta}.xlsx")
        origen = os.path.join(script_directory, DEPENDENCIAS_BD_Y_PLANTILLAS,plantilla)
        shutil.copy(origen, destino)
    # Copiar archivos adicionales
    shutil.copy(os.path.join(script_directory,"SCR", "second.png"), ruta_nueva_carpeta)

    #Mover pdf a la carpeta
    if rutaPdf:
        shutil.copy(rutaPdf,ruta_nueva_carpeta)
        return True
    else:
        return False

# Función para encontrar un archivo PDF que comience con 'FSC' y extraer un nombre comercial basado en un código presente en el nombre del archivo
def encontrar_pdf_y_extraer_nombre(archivo):

    nombres_Comercial = {
        'CG': 'CAROLINA GAITAN',
        'ER': 'ESTEBAN RAMIREZ',
        'JO': 'JIMMY ORTIZ',
        'NF': 'NELIS FABREGAS',
        'JC': 'JEIMY CADENA',
        'JR': 'JENIFER RIOS',
        'RB': 'ROCÍO BARÓN',
        'DC': 'DANIELA CRISTANCHO',
        'MB': 'MARIA MERCEDES BURGOS',
        'CC': 'DIANA CORREA',
        'LC': 'LUIS CAÑÓN'
    }
    
    archivo=archivo.split('/')[-1]
    
    if archivo.startswith('FSC'):
        return nombres_Comercial[archivo.split('-')[1]]  # Extracts the code from the filename and returns the corresponding commercial name.
    return None

# Función para buscar en un DataFrame de pandas las referencias proporcionadas y notificar cuáles se encontraron y cuáles no
def extraer_referencias_de_base_de_datos(referencias):
    """
    Busca en todos los dataFrames (PHYWE+EUROMEX) 
    Devuelve las referencias encontradas y notifica de las no encontradas.
    """
    df1=baseDeDatos_dataFrame('PHYWE')
    df2=baseDeDatos_dataFrame('EUROMEX')
    df = pd.concat([df1, df2], axis=0)
    
    referencias_existentes = []
    referencias_no_encontradas = []

    for ref in referencias:
        if ref in df.index:
            referencias_existentes.append(ref)  # Adds the reference to the list of existing ones if found in the DataFrame's index.
        else:
            referencias_no_encontradas.append(ref)  # Adds the reference to the list of not found in the opposite case.

    return df.loc[referencias_existentes]  # Returns the rows of the DataFrame corresponding to the existing references.

# Función para manejar el llenado de un archivo SP con información específica
def manejar_SP(dataUI, df_SP, cantidades,marcaDestino):
    """
    Llena el encabezado del SP. Extrae del nombre del archivo el nombre de universidad y consecutivo,
    los cuales retorna para posterior uso.
    Extrae referencias de la base de datos y la llena en la tabla del SP.
    La única columna que queda por llenar es CANTIDADES, las cuales las pide al Usuario.
    Por último, llena la segunda hoja del excel SP con datos del usuario también
    """
    
    nombre_carpeta=dataUI['Carpeta']
    rutaCarpeta=os.path.join(script_directory,marcaDestino,nombre_carpeta)

    archivos = os.listdir(rutaCarpeta)    
    # Encuentra el archivo que comienza con 'OFERTA' y construye la ruta completa al archivo encontrado
    archivo_of = next((archivo for archivo in archivos if archivo.startswith('OFERTA')), None)
    rutaOF = os.path.join(rutaCarpeta, archivo_of)

    # Encuentra el archivo que comienza con 'SP' y construye la ruta completa al archivo encontrado
    archivo_SP = next((archivo for archivo in archivos if archivo.startswith('SP')), None)
    rutaSP = os.path.join(rutaCarpeta, archivo_SP)

    wb_sp = openpyxl.load_workbook(rutaSP)
    hoja_SP = wb_sp.worksheets[0]

    estampillas=float(dataUI['Estampillas'])/100
    imprevistos=float(dataUI['Imprevistos'])/100
    ciudad = dataUI['Ciudad']
    institucion=dataUI['Institucion']    
    consecutivo= dataUI['Consecutivo']
    comercial = dataUI['Comercial']

    hoja_SP['E2'] = estampillas
    hoja_SP['E3'] = imprevistos
    hoja_SP['E14'] = ciudad
    hoja_SP['E15'] = dataUI['Trimestre']
    hoja_SP['E16'] = dataUI['Presupuesto']
    hoja_SP['E6'] = institucion
    hoja_SP['E17']= dataUI['Moneda']


    if comercial in ['JIMMY ORTIZ', 'JEIMY CADENA','LUIS CAÑÓN', 'MARIA MERCEDES BURGOS']:
        hoja_SP['E8'] = comercial  # Updates cell E8 with the commercial name if it's Ortiz or Cadena.
    else:
        hoja_SP['E9'] = comercial

    # Iterates through the rows of the DataFrame and inserts them into the Excel sheet.
    for r_idx, row in enumerate(dataframe_to_rows(df_SP, index=True, header=False), 19):
        if r_idx == 19:
            continue  # Skips the first row (DataFrame indices).
        for c_idx, value in enumerate(row, 1):
            hoja_SP.cell(row=r_idx, column=c_idx, value=value)  # Inserts each value into the corresponding cell.

    
    if isinstance(cantidades, list):

        for j, valor in enumerate(cantidades, 20):
            hoja_SP.cell(row=j, column=9, value=int(valor))


    
    #Ahora llena la segunda hoja del excel
    hoja_SP_gastos=wb_sp.worksheets[1]

    if dataUI['Dias']=='' or dataUI['Profesionales']=='':
        pass
    else:
        hoja_SP_gastos['D6']=int(dataUI['Dias'])
        hoja_SP_gastos['D7']=int(dataUI['Profesionales'])


    wb_sp.save(rutaSP)  # Saves the changes made in the Excel file.

    #Ahora Manejar la Oferta
    wb_OF = openpyxl.load_workbook(rutaOF)  # Loads the Excel workbook.
    hoja_destino= wb_OF.worksheets[0]

    #Llena y formatea el encabezado
    celda_cliente=hoja_destino['B17:D17']
    celda_cliente[0][0].value= f"NOMBRE DE CLIENTE: {institucion}"

    celda_consecutivo=hoja_destino['E17:F17']
    celda_consecutivo[0][0].value= f"OFERTA N°: {consecutivo}"
    
    celda_ciudad=hoja_destino['G17:H17']
    celda_ciudad[0][0].value= f"CIUDAD: {ciudad}"
    
    celda_comercial=hoja_destino['B18:D18']
    celda_comercial[0][0].value= f"COMERCIAL: {comercial}"

    celda_estampillas=hoja_destino['K17:L17']
    celda_estampillas[0][0].value= f"% ESTAMPILLAS: {estampillas} %"

    celda_imprevistos=hoja_destino['K18:L18']
    celda_imprevistos[0][0].value= f"% IMPREVISTOS: {imprevistos} %"


    # Aplicar formato (negrita y Arial Narrow) a las celdas
    for celda in ['B17', 'G17', 'B18', 'K17', 'K18', 'E17']:
        hoja_destino[celda].font = Font(bold=True, name='Arial Narrow')

    # Colocar las respuestas en las celdas correspondientes
    hoja_destino['F18'] = dataUI['Requerimiento']
    hoja_destino['H18'] = dataUI['Tipo']
    hoja_destino['J18'] = dataUI['Canal']

    img = Image(os.path.join(rutaCarpeta, 'second.png'))
    hoja_destino.add_image(img, 'B1')
    wb_OF.save(rutaOF)

    os.remove(os.path.join(rutaCarpeta, 'second.png'))

# Función para crear un archivo CSV de cotización a partir de un archivo SP
def crear_csv_cot(rutaCarpeta):
    """
    Crea el csv de cotización de PHYWE a partir del SP. 
    Se utiliza Pandas para extraer del SP y escribir en el csv
    """
    archivos = os.listdir(rutaCarpeta) 
    archivo_SP = next((archivo for archivo in archivos if archivo.startswith('SP')), None)
    ruta_del_archivo=os.path.join(rutaCarpeta,archivo_SP)
    df = pd.read_excel(ruta_del_archivo, skiprows=18, sheet_name='SOLICITUD')
    df_soloPhywe = df[(df['PROVEEDOR'] == 'PHYWE')]
    nuevoDF = pd.DataFrame({
        'Ref': df_soloPhywe['REFERENCIA'],
        'Qty': df_soloPhywe['CANTIDAD']
    }).replace(r'\s+', '', regex=True)

    output_file = os.path.join(rutaCarpeta, 'PHYWEQUOTE.csv')
    nuevoDF.to_csv(output_file, sep=';', index=False)

def gastos_de_viaje():
    ruta_gastos= os.path.join(script_directory,DEPENDENCIAS_BD_Y_PLANTILLAS, 'SP.xlsx')
    data=pd.read_excel(ruta_gastos,index_col=0,sheet_name='TARIFAS')
    return data.index.to_list()

def baseDeDatos_dataFrame(marca):
    ruta_BasedeDatos = os.path.join(script_directory,DEPENDENCIAS_BD_Y_PLANTILLAS, ARCHIVO_ESPECIFICACIONES_EXCEL)
    df = pd.read_excel(ruta_BasedeDatos, index_col=0,sheet_name=marca)
    df = df.drop(columns=['FILTRADO'])
    return df  

def nombres_de_basedeDatos(marca):
    df=baseDeDatos_dataFrame(marca)
    nombres=df.loc[:,'DESCRIPCION']
    return nombres.items()

def obtener_nuevo_consecutivo(prefijo,marcaDestino):
    """
    Dado un prefijo, esta función busca en el directorio actual por carpetas
    que comiencen con ese prefijo, extrae el número consecutivo, y devuelve
    el último consecutivo encontrado más uno.
    """
    consecutivo_maximo = 0
    pattern = re.compile(rf"^{prefijo}\s+(\d+)-")
    carpeta_real=os.path.join(script_directory,marcaDestino)

    for carpeta in os.listdir(carpeta_real):
        full_path = os.path.join(carpeta_real, carpeta)
        if os.path.isdir(full_path):
            match = pattern.match(carpeta)
            if match:
                consecutivo_actual = int(match.group(1))
                consecutivo_maximo = max(consecutivo_maximo, consecutivo_actual)    
    return consecutivo_maximo + 1


